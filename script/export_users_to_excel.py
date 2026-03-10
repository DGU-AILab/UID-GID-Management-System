#!/usr/bin/env python3
"""Export active/deleted user data from LAB/FARM databases to Excel and Google Sheets."""

import argparse
import os
import sys
from datetime import datetime

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

SPREADSHEET_ID = '1U3-YidZrxNHH4mEbq6-MaxZqsUWJ6HZn2GV9flwIwPY'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
VALID_DOMAINS = ('LAB', 'FARM')


def parse_args():
    parser = argparse.ArgumentParser(
        description='사용자 정보를 MySQL 데이터베이스에서 추출하여 Excel/Google Sheets로 저장합니다.'
    )
    parser.add_argument(
        '--domains',
        help='조회할 도메인 목록 (예: LAB,FARM). 기본값은 설정 파일의 EXPORT_DOMAINS 또는 SERVER_DOMAIN.',
    )
    return parser.parse_args()


def resolve_db_config_path():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    config_path = os.path.join(project_root, 'config', 'db_config.local.env')
    if os.path.exists(config_path):
        return config_path

    raise FileNotFoundError(
        'config/db_config.local.env not found. Copy config/db_config.example.env to config/db_config.local.env first.'
    )


def resolve_google_client_credentials_path(project_root):
    return os.path.join(project_root, 'config', 'google-client.local.json')


def load_raw_config():
    config = {}
    with open(resolve_db_config_path(), 'r', encoding='utf-8') as handle:
        for line in handle:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                config[key.strip()] = value.strip()
    return config


def normalize_domain(domain_name):
    normalized = domain_name.strip().upper()
    if normalized not in VALID_DOMAINS:
        raise ValueError(f'Unsupported domain: {domain_name}')
    return normalized


def resolve_domains(raw_config, raw_domains):
    domains_value = raw_domains or raw_config.get('EXPORT_DOMAINS') or raw_config.get('SERVER_DOMAIN') or 'LAB'
    domains = [normalize_domain(item) for item in domains_value.split(',') if item.strip()]
    if not domains:
        raise ValueError('At least one domain is required.')
    return list(dict.fromkeys(domains))


def resolve_db_host_for_domain(raw_config, domain_name):
    specific_key = f'{domain_name}_DB_HOST'
    if raw_config.get(specific_key):
        return raw_config[specific_key]
    if raw_config.get('DB_HOST'):
        return raw_config['DB_HOST']
    raise ValueError(f'{specific_key} or DB_HOST must be configured.')


def build_db_config(raw_config, domain_name):
    return {
        'host': resolve_db_host_for_domain(raw_config, domain_name),
        'port': int(raw_config['DB_PORT']),
        'user': raw_config['DB_USER'],
        'password': raw_config['DB_PASSWORD'],
        'database': raw_config['DB_NAME'],
        'charset': raw_config['DB_CHARSET'],
    }


def get_user_data(db_config, existing_only=True):
    try:
        connection = pymysql.connect(**db_config)
        cursor = connection.cursor()

        query = """
        SELECT
            '' AS '존재여부',
            u.name AS '이름',
            u.ubuntu_username AS '로그인 아이디',
            g.ubuntu_groupname AS '그룹명',
            dc.server_id AS '배정 서버',
            u.ubuntu_uid AS 'UID',
            u.ubuntu_gid AS 'GID',
            GROUP_CONCAT(DISTINCT up.port_number ORDER BY up.port_number SEPARATOR ', ') AS '포트 번호',
            DATE_FORMAT(dc.expiring_at, '%%Y-%%m-%%d') AS '서버 사용 완료 예정일',
            DATE_FORMAT(DATE_ADD(dc.expiring_at, INTERVAL 15 DAY), '%%Y-%%m-%%d') AS '스토리지 삭제 예정일',
            dc.created_by AS '컨테이너 생성자',
            DATE_FORMAT(dc.created_at, '%%Y-%%m-%%d') AS '컨테이너 생성일자',
            CONCAT(dc.image, ':', dc.image_version) AS 'docker image version',
            CONCAT(dc.container_id, ' | ', dc.container_name) AS '컨테이너 명',
            COALESCE(u.email, '') AS 'E-mail',
            COALESCE(u.phone, '') AS '전화번호',
            '' AS '사용여부',
            u.note AS '비고'
        FROM user u
        LEFT JOIN `group` g ON u.ubuntu_gid = g.ubuntu_gid
        JOIN docker_container dc ON u.id = dc.user_id
        LEFT JOIN used_ports up ON dc.id = up.docker_container_record_id
        WHERE dc.existing = %s
          AND NULLIF(TRIM(dc.server_id), '') IS NOT NULL
        GROUP BY u.id, dc.id
        ORDER BY dc.server_id ASC, u.name ASC;
        """

        cursor.execute(query, [1 if existing_only else 0])
        columns = [desc[0] for desc in cursor.description]
        rows = cursor.fetchall()
        cursor.close()
        connection.close()
        return columns, rows
    except pymysql.Error as exc:
        print(f'데이터베이스 오류: {exc}')
        sys.exit(1)


def create_excel_sheet(workbook, sheet_name, columns, data):
    if len(workbook.sheetnames) == 1 and workbook.active.title == 'Sheet':
        worksheet = workbook.active
        worksheet.title = sheet_name
    else:
        worksheet = workbook.create_sheet(title=sheet_name)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = column_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical='center')

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    return worksheet


def get_google_sheets_service(credentials_path):
    try:
        credentials = service_account.Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
        return build('sheets', 'v4', credentials=credentials)
    except Exception as exc:
        print(f'Google Sheets 인증 오류: {exc}')
        return None


def ensure_sheet_exists(service, sheet_name):
    try:
        spreadsheet = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID).execute()
        for sheet in spreadsheet.get('sheets', []):
            if sheet['properties']['title'] == sheet_name:
                print(f"✓ '{sheet_name}' 시트가 이미 존재합니다.")
                return sheet['properties']['sheetId']

        response = service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={'requests': [{'addSheet': {'properties': {'title': sheet_name}}}]},
        ).execute()
        sheet_id = response['replies'][0]['addSheet']['properties']['sheetId']
        print(f"✓ '{sheet_name}' 시트가 생성되었습니다.")
        return sheet_id
    except HttpError as exc:
        print(f'시트 확인/생성 오류: {exc}')
        return None


def format_header(service, num_columns, sheet_id):
    try:
        requests = [{
            'repeatCell': {
                'range': {
                    'sheetId': sheet_id,
                    'startRowIndex': 0,
                    'endRowIndex': 1,
                    'startColumnIndex': 0,
                    'endColumnIndex': num_columns,
                },
                'cell': {
                    'userEnteredFormat': {
                        'backgroundColor': {'red': 0.267, 'green': 0.447, 'blue': 0.769},
                        'textFormat': {
                            'bold': True,
                            'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0},
                        },
                        'horizontalAlignment': 'CENTER',
                        'verticalAlignment': 'MIDDLE',
                    }
                },
                'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
            }
        }]
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={'requests': requests},
        ).execute()
        print('✓ Google Sheets 헤더 서식 적용 완료')
    except HttpError as exc:
        print(f'헤더 서식 적용 오류: {exc}')


def update_google_sheet(service, columns, data, sheet_name):
    try:
        sheet_id = ensure_sheet_exists(service, sheet_name)
        if sheet_id is None:
            return False

        values = [list(columns)]
        values.extend([['' if value is None else str(value) for value in row] for row in data])

        service.spreadsheets().values().clear(
            spreadsheetId=SPREADSHEET_ID,
            range=f'{sheet_name}!A:Z',
        ).execute()

        result = service.spreadsheets().values().update(
            spreadsheetId=SPREADSHEET_ID,
            range=f'{sheet_name}!A1',
            valueInputOption='RAW',
            body={'values': values},
        ).execute()

        format_header(service, len(columns), sheet_id)
        updated_cells = result.get('updatedCells', 0)
        print(f"✓ Google Sheets '{sheet_name}' 시트 업데이트 완료: {updated_cells}개의 셀이 업데이트되었습니다.")
        return True
    except HttpError as exc:
        print(f'Google Sheets API 오류: {exc}')
        return False
    except Exception as exc:
        print(f'Google Sheets 업데이트 오류: {exc}')
        return False


def main():
    args = parse_args()
    raw_config = load_raw_config()
    domains = resolve_domains(raw_config, args.domains)

    print('=' * 60)
    print('사용자 정보 엑셀 추출 스크립트')
    print('=' * 60)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    export_dir = os.path.join(project_root, 'excel_exports')
    os.makedirs(export_dir, exist_ok=True)

    today = datetime.now().strftime('%Y-%m-%d')
    filename = os.path.join(export_dir, f'user_export_{today}.xlsx')

    print(f'\n저장 디렉토리: {export_dir}')
    print(f'대상 도메인: {", ".join(domains)}')

    workbook = Workbook()

    for domain in domains:
        db_config = build_db_config(raw_config, domain)
        print(f'\n[{domain}] 데이터베이스 연결 중... ({db_config["host"]}:{db_config["port"]})')

        print(f'[{domain}] 활성 사용자 데이터 조회 중...')
        active_columns, active_data = get_user_data(db_config, existing_only=True)
        print(f"✓ [{domain}] 활성 사용자 데이터 조회 완료: {len(active_data)}개의 레코드")

        print(f'[{domain}] 삭제된 사용자 데이터 조회 중...')
        deleted_columns, deleted_data = get_user_data(db_config, existing_only=False)
        print(f"✓ [{domain}] 삭제된 사용자 데이터 조회 완료: {len(deleted_data)}개의 레코드")

        create_excel_sheet(workbook, domain, active_columns, active_data)
        create_excel_sheet(workbook, f'{domain}(deleted)', deleted_columns, deleted_data)

    print(f'\n엑셀 파일 생성 중: {os.path.basename(filename)}')
    workbook.save(filename)
    print(f'✓ 엑셀 파일이 성공적으로 생성되었습니다: {filename}')

    print('\nGoogle Sheets 업데이트 중...')
    credentials_path = resolve_google_client_credentials_path(project_root)
    if os.path.exists(credentials_path):
        service = get_google_sheets_service(credentials_path)
        if service:
            for domain in domains:
                db_config = build_db_config(raw_config, domain)
                active_columns, active_data = get_user_data(db_config, existing_only=True)
                deleted_columns, deleted_data = get_user_data(db_config, existing_only=False)
                update_google_sheet(service, active_columns, active_data, sheet_name=domain)
                update_google_sheet(service, deleted_columns, deleted_data, sheet_name=f'{domain}(deleted)')
        else:
            print('⚠ Google Sheets 서비스 연결 실패')
    else:
        print(f'⚠ 인증 파일을 찾을 수 없습니다: {credentials_path}')

    print('\n' + '=' * 60)
    print('작업 완료!')
    print('=' * 60)


if __name__ == '__main__':
    main()
